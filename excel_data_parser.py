'''The Excel Data Parser is not called anywhere in the main program. This was run
separately to extract the data and create CSV files. The main program only opens
the CSV files and gathers data.'''

import openpyxl
import re

def parse_packages():
    '''Function parses the excel file and save its contents into csv'''
    data = []
    file_location = "./resources/WGUPSPackageFile.xlsx"

    packages_wb = openpyxl.load_workbook(file_location, data_only=True)
    packages_sheet = packages_wb.active

    for row in range(packages_sheet.max_row):
        data.append([])
        for col in range(packages_sheet.max_column):
            data[row].append(str(packages_sheet.cell(row+1, col+1).internal_value))
        data[row] = ','.join(data[row])
        print(data[row])

    # Write to file after appending new fields for status (i.e. At hub, in transit, etc.) and delivery time
    with open("./resources/packages.txt", "w") as packages_file:
        temp_string = ''
        for i in range(len(data)):
            temp_string = str(data[i] + ', At Hub' + ',Not Delivered Yet' + '\n')
            packages_file.write(temp_string)


def parse_distances():
    '''Function parses the distance table excel file and save contents to csv'''

    file_location = './resources/WGUPSDistanceTable.xlsx'
    distances_wb = openpyxl.load_workbook(file_location, data_only=True)
    distances_sheet = distances_wb.active
    data = [] #create empty list to hold data

    for row in range(2, distances_sheet.max_row+1):
        for col in range(2, distances_sheet.max_column):

            #if statement prevents duplicate data from symmetrical excel table (also entries of the same spot, distance = 0)
            if distances_sheet.cell(row, col).internal_value == None:
                break

            # Uses regex to remove zip codes from the addresses in column 1
            # col1 is the first column value in the CSV file
            col1 = re.sub(r'\([^)]*\)', '', distances_sheet.cell(row, 1).internal_value) #regex removes zipcodes

            # Uses regex search to remove the first and third lines of the excel column address values in row 1
            # Then it removes any leftover commas from the long-form address and strips white space
            # col2 is the second column of the final CSV file
            col2_matches = re.search(r'\n.*', distances_sheet.cell(1, col).internal_value)
            col2 = col2_matches.group(0)
            col2 = col2.replace(',', '')
            col2 = col2.strip()

            # Locates the cell value based on the two addresses from col1 and col2
            # Must be converted to str because cell value will return a float
            col3 = str(distances_sheet.cell(row, col).internal_value)

            # Creates the final string for data insertion
            # First, concatenates and adds commas for CSV
            # Then, removes any leftover newlines and trailing/leading whitespaces
            concatenated_cols = col1 + ',' + col2 + ',' + col3
            concatenated_cols = concatenated_cols.replace("\n", '') #strip the numerous newlines
            concatenated_cols = concatenated_cols.strip() #strip trailing and leading whitespaces

            data.append(concatenated_cols)

    # Write to CSV file and place each string of data onto a new line
    with open("./resources/distances.txt", "w") as distances_file:
        temp_string = ''
        for i in range(len(data)):
            temp_string = str(data[i] + '\n')
            distances_file.write(temp_string)

# Run functions to parse excel sheets
parse_packages()
parse_distances()